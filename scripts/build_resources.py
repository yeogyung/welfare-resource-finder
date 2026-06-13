from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT.parent

SOURCES = {
    "top10": WORKSPACE / "배경지식/00_원본자료/다운로드_첨부자료/찾아봄_서울노인복지_TOP10_원본형식_정제최종 (1).xlsx",
    "gov24": WORKSPACE / "배경지식/00_원본자료/다운로드_첨부자료/정부24_복지서비스_API수집_wlsWk 최종라벨링 2.xlsx",
    "bokjiro": WORKSPACE / "배경지식/00_원본자료/다운로드_첨부자료/복지로_API 수집_v3 (1).xlsx",
    "gwangjin": WORKSPACE / "배경지식/00_원본자료/다운로드_첨부자료/광진구_민간복지자원_라벨링_60개.xlsx",
}

CATEGORY_KEYWORDS = {
    "emergency": ["응급", "위기", "긴급", "쓰러", "안전", "학대", "자살", "고독사"],
    "care": ["돌봄", "맞춤돌봄", "재가", "방문", "식사", "도시락", "반찬", "생활지원", "요양"],
    "health": ["건강", "의료", "병원", "검진", "수술", "치매", "무릎", "눈", "안검진", "보건"],
    "money": ["경제", "생활비", "연금", "현금", "수당", "의료비", "지원금", "일자리", "취업"],
    "housing": ["주거", "집", "난방", "에너지", "환경개선", "수리", "담보", "임대"],
    "learning": ["교육", "디지털", "스마트폰", "컴퓨터", "금융", "문해", "배움", "강좌"],
    "culture": ["문화", "여가", "공연", "전시", "모임", "사회참여", "체조", "운동", "서예", "무용"],
}

CATEGORY_LABELS = {
    "emergency": "응급·안전",
    "care": "돌봄·생활",
    "health": "건강·의료",
    "money": "경제·일자리",
    "housing": "주거·환경",
    "learning": "디지털·배움",
    "culture": "문화·여가",
    "general": "기타 지원",
}


def clean(value: Any, limit: int | None = None) -> str:
    if value is None:
        return ""
    text = str(value).replace("_x000D_", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if limit and len(text) > limit:
        return text[: limit - 1].rstrip() + "…"
    return text


def rows_from_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb[sheet_name]
    headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        if clean(item.get("서비스명")):
            rows.append(item)
    return rows


def category_for(*parts: str) -> str:
    text = " ".join(parts)
    for key, keywords in CATEGORY_KEYWORDS.items():
        if any(k in text for k in keywords):
            return key
    return "general"


def normalize(row: dict[str, Any], source: str, source_label: str, priority: int) -> dict[str, Any]:
    name = clean(row.get("서비스명"), 120)
    desc = clean(row.get("지원내용") or row.get("servDgst"), 280)
    method = clean(row.get("신청방법"), 260)
    target = clean(row.get("대상") or row.get("대상태그") or row.get("대상 태그") or row.get("trgterIndvdlArray"), 220)
    category_raw = clean(row.get("대분류") or row.get("유형") or row.get("기관유형") or row.get("intrsThemaArray"), 80)
    situation = clean(row.get("상황키워드") or row.get("상황 키워드"), 120)
    region = clean(row.get("지역구") or row.get("지역") or "", 40)
    link = clean(row.get("링크") or row.get("servDtlLink"))
    valid = clean(row.get("유효일") or row.get("유효일(지원주기)") or row.get("종료일") or row.get("sprtCycNm"), 80)
    org = clean(row.get("기관명") or row.get("jurOrgNm") or row.get("jurMnofNm"), 80)
    contact = ""
    contact_match = re.search(r"(?:0\d{1,2}-\d{3,4}-\d{4}|1[235]\d{2}|129|1688-\d{4}|1544-\d{4}|1577-\d{4}|1670-\d{4})", " ".join([method, target, desc]))
    if contact_match:
        contact = contact_match.group(0)
    elif source in {"gov24", "bokjiro"}:
        contact = "129"
    category = category_for(category_raw, situation, name, desc, target)
    if source == "gwangjin" and not region:
        region = "광진구"
    if source in {"gov24", "bokjiro", "top10"} and not region:
        region = "전국/서울 가능"
    if not valid:
        valid = "상세 확인 필요"
    requires_check = not link or "문의" in link or not method
    return {
        "name": name,
        "source": source,
        "sourceLabel": source_label,
        "priority": priority,
        "category": category,
        "categoryLabel": CATEGORY_LABELS[category],
        "categoryRaw": category_raw,
        "situation": situation,
        "target": target or "대상 확인 필요",
        "description": desc or "상세 내용 확인 필요",
        "method": method or "공식 홈페이지 또는 129 문의로 확인 필요",
        "period": valid,
        "region": region,
        "organization": org,
        "contact": contact or "129",
        "url": link,
        "requiresCheck": requires_check,
        "searchText": " ".join([name, category_raw, situation, target, desc, method, region, org]),
    }


def build() -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in rows_from_sheet(SOURCES["top10"], "서울_TOP10_원본형식"):
        items.append(normalize(row, "top10", "찾아봄 TOP10", 100))
    for row in rows_from_sheet(SOURCES["gwangjin"], "광진구_민간복지_라벨링"):
        items.append(normalize(row, "gwangjin", "광진구 민간복지", 90))
    for row in rows_from_sheet(SOURCES["bokjiro"], "복지로_정리본"):
        items.append(normalize(row, "bokjiro", "복지로", 70))
    for row in rows_from_sheet(SOURCES["gov24"], "정부24_라벨링"):
        items.append(normalize(row, "gov24", "정부24", 60))

    deduped: dict[str, dict[str, Any]] = {}
    for item in items:
        key = (item["name"] + "|" + item.get("url", "")).strip()
        if key not in deduped or item["priority"] > deduped[key]["priority"]:
            deduped[key] = item

    output = []
    for idx, item in enumerate(sorted(deduped.values(), key=lambda x: (-x["priority"], x["name"])), 1):
        item["id"] = f"wr-{idx:04d}"
        output.append(item)
    return output


def main() -> None:
    resources = build()
    out = ROOT / "public/data/welfare-resources.json"
    out.write_text(json.dumps(resources, ensure_ascii=False, indent=2), encoding="utf-8")
    stats: dict[str, Any] = {
        "total": len(resources),
        "bySource": {},
        "byCategory": {},
    }
    for item in resources:
        stats["bySource"][item["sourceLabel"]] = stats["bySource"].get(item["sourceLabel"], 0) + 1
        stats["byCategory"][item["categoryLabel"]] = stats["byCategory"].get(item["categoryLabel"], 0) + 1
    (ROOT / "public/data/resource-stats.json").write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
